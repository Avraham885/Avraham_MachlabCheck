# ==== UI SHELL (Single Instance) ====
import os
import streamlit as st

def load_password():
    # 1) Streamlit secrets (Cloud או מקומי)
    try:
        return st.secrets["auth"]["password"]
    except Exception:
        pass
    # 2) משתנה סביבה כגיבוי (אופציונלי)
    return os.getenv("AUTH_PASSWORD")

PASSWORD = load_password()
st.title("🔐 הזדהות נדרשת")

if not PASSWORD:
    st.error("הסיסמה אינה מוגדרת. הגדירו .streamlit/secrets.toml מקומי או AUTH_PASSWORD בסביבה.")
    st.stop()

password = st.text_input("הזן/י סיסמה לגישה:", type="password")
if password != PASSWORD:
    st.warning("יש להזין סיסמה נכונה כדי להמשיך.")
    st.stop()

st.success("גישה מאושרת ✅")


st.markdown("""
<style>
:root{
  --bg: #0ea5e9; --bg2:#38bdf8; --ink:#0f172a; --sub:#475569;
  --muted:#f1f5f9; --border:#e2e8f0; --accent:#2563eb; --accent-ink:#ffffff; --radius:14px;
}
html, body, [class*="css"] { direction: rtl; text-align: right; }
.block-container { padding-top: 8px; padding-bottom: 18px; }
section.main h1 { display: none; } /* hide any existing st.title to prevent duplicate headline */

.hero { background: linear-gradient(90deg, var(--bg), var(--bg2)); color: white; padding: 22px 24px; border-radius: 20px; margin: 6px 0 16px 0; box-shadow: 0 8px 18px rgba(2,132,199,0.18); }
.hero h1 { margin: 0; font-size: 30px; line-height: 1.15; font-weight: 800; letter-spacing: .2px; }
.hero p { margin: 6px 0 0 0; font-size: 15px; opacity: .95; }

.tip-banner { background: var(--muted); border: 1px solid var(--border); padding: .55rem .8rem; border-radius: 12px; margin: 8px 0 16px 0; font-size: 13px; color: var(--ink); }
.section-title{ font-size: 18px; font-weight: 700; color: var(--ink); margin: 8px 0 6px 0; }
hr { border: none; border-top: 1px solid var(--border); margin: 12px 0 14px 0; }

[data-testid="stFileUploaderDropzone"]{ border-radius: 14px; background: #f8fbff; border: 1px dashed #b6cffc; }
.stButton>button, .stDownloadButton>button{ border-radius: var(--radius); padding: .7rem 1.05rem; font-weight: 700; border: 1px solid var(--accent); background: var(--accent); color: var(--accent-ink); box-shadow: 0 4px 10px rgba(37,99,235,.18); }
.stDownloadButton>button:hover, .stButton>button:hover{ opacity: .96; }

.footer{ background: #fbfdff; border-top: 1px solid var(--border); padding: 10px 12px; border-radius: 14px; margin-top: 18px; color: #475569; font-size: 12px; text-align:center; }
.footer b { color:#0f172a; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero">
  <h1>בדיקת דוח התחשבנות מחלב אלכל</h1>
  <p>בדיקת דוח התחשבנות מחלב ואלכל. מעלים את הקבצים, לוחצים ‘בדיקה’, ומקבלים קובץ מעודכן להורדה — פשוט.</p>
</div>
<div class="tip-banner"><b>טיפ:</b> שמרו על שמות גיליונות זהים לתבנית כדי לקבל תוצאות מדויקות.</div>
""", unsafe_allow_html=True)
# ==== END SHELL ====

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Machlab – נוסחאות מוגבלות לפי 'הזמנות לבדיקה'", layout="wide")

# ----- RTL UI -----
st.markdown("""
<style>
html, body, [class*="css"]  { direction: rtl; text-align: right; }
</style>
""", unsafe_allow_html=True)

st.title("העלאת קבצים")
st.caption("מעלה שני קבצים (פעילות + 'הובלות אלכל כללי'), מעתיק גיליון פנימי, מוסיף עמודות חסרות ומזריק נוסחאות. "
           "VLOOKUP משתמש ב-MATCH דינמי לעמודה 'סופק'. כל ההזרקות נעצרות בשורה האחרונה של 'הזמנות לבדיקה'. "
           "הסקריפט מוחל גם על 'הובלה לבית לקוח' וגם על 'הובלה לסוחר'.")

# ----- Constants -----
REQUIRED_MAIN_SHEET_1     = "הובלה לבית לקוח"
REQUIRED_MAIN_SHEET_2     = "הובלה לסוחר"
REQUIRED_INTERNAL_SHEET   = "הובלות אלכל כללי"

COL_PURCHASE_SRC  = "הז. רכש (לקוח)"
COL_RAKHASH       = "רכש"
COL_MAKAT         = "מק'ט"
COL_MAKAT_CLEAN   = "מקט ללא פגומים"
COL_ORDER_CHECK   = "הזמנות לבדיקה"
COL_QTY           = "כמות"
COL_QTY_CHECK     = "בדיקת כמות"
COL_PRICE_AFTER   = "מחירון מחלב לאחר בדיקה"
COL_DUP_JULY      = "כפילויות חודש קודם"
COL_MANUAL        = "מעבר ידני"
COL_APPROVAL      = "אישור סופי"
COL_NOTES         = "הערות"
COL_TOTAL_PAY     = "סה\"כ לתשלום"
COL_DIFF_ROW      = "פער לפי שורה"
COL_TOTAL         = "סה\"כ"  # עמודת סיכום קיימת אם יש

# עמודות נוספות להוספה אם חסרות (ללא נוסחאות, מלבד אלה שמוגדר להן נוסחה)
EXTRA_COLUMNS = [
    COL_PRICE_AFTER,
    COL_DUP_JULY,
    COL_QTY_CHECK,        # עם נוסחת VLOOKUP+MATCH
    COL_MANUAL,
    COL_APPROVAL,         # IF/OR/AND
    COL_NOTES,
    COL_TOTAL_PAY,        # IF(...*ABS(...))
    COL_DIFF_ROW,         # סה"כ - סה"כ לתשלום
]

# ----- Helpers -----
def find_sheet_name(sheetnames, target):
    if target in sheetnames:
        return target
    if (target + " ") in sheetnames:
        return target + " "
    for s in sheetnames:
        if s.strip() == target.strip():
            return s
    return None

def find_col(ws, name):
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if (val or "").strip() == name:
            return c
    return None

def ensure_column(ws, name):
    """Locate a column by header; if missing, append it at the end and return the index."""
    idx = find_col(ws, name)
    if idx is not None:
        return idx, False
    new_idx = ws.max_column + 1 if ws.max_column else 1
    ws.cell(row=1, column=new_idx, value=name)
    return new_idx, True

def copy_dataframe_to_sheet(df: pd.DataFrame, wb: Workbook, sheet_name: str):
    # overwrite if exists
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    # RTL view only (do not reverse column order)
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass
    # headers
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=str(col))
    # rows
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

def last_nonempty_row(ws, col_letter, start_row=2):
    max_row = ws.max_row
    for r in range(max_row, start_row - 1, -1):
        v = ws[f"{col_letter}{r}"].value
        if v is not None and str(v).strip() != "":
            return r
    return start_row - 1

# ----- UI -----
col1, col2 = st.columns(2)
with col1:
    activity_file = st.file_uploader("קובץ פעילות (Excel) – בדרך כלל 'פעילות אלכל חודש שנה' (.xlsx)", type=["xlsx"], key="activity")
with col2:
    internal_file = st.file_uploader("קובץ 'הובלות אלכל כללי' (Excel) (.xlsx)", type=["xlsx"], key="internal")

if activity_file and internal_file:
    try:
        # 1) Load activity workbook
        wb = load_workbook(activity_file)

        # 2) Copy internal sheet as a new sheet in the same workbook (once)
        xl_internal = pd.ExcelFile(internal_file)
        internal_sheet_name = find_sheet_name(xl_internal.sheet_names, REQUIRED_INTERNAL_SHEET)
        if not internal_sheet_name:
            st.error(f"לא נמצא גיליון בשם '{REQUIRED_INTERNAL_SHEET}' בקובץ 'הובלות אלכל כללי'.")
            st.stop()
        df_internal = xl_internal.parse(internal_sheet_name)
        copy_dataframe_to_sheet(df_internal, wb, REQUIRED_INTERNAL_SHEET)

        TARGET_SHEETS = [REQUIRED_MAIN_SHEET_1, REQUIRED_MAIN_SHEET_2]
        overall_msgs = []

        # 3) Process each target sheet with the exact same injections
        for target_sheet in TARGET_SHEETS:
            main_sheet_name = find_sheet_name(wb.sheetnames, target_sheet)
            if not main_sheet_name:
                st.warning(f"לא נמצא גיליון בשם '{target_sheet}' בקובץ הפעילות – מדלג/ה.")
                continue
            ws = wb[main_sheet_name]

            # Ensure required and extra columns exist (add at end if missing)
            col_purchase_src = ensure_column(ws, COL_PURCHASE_SRC)[0]
            col_rakhash, _   = ensure_column(ws, COL_RAKHASH)
            col_makat, _     = ensure_column(ws, COL_MAKAT)
            col_clean, _     = ensure_column(ws, COL_MAKAT_CLEAN)
            col_order, _     = ensure_column(ws, COL_ORDER_CHECK)
            col_qty, _       = ensure_column(ws, COL_QTY)
            col_qty_check,_  = ensure_column(ws, COL_QTY_CHECK)
            col_price,_      = ensure_column(ws, COL_PRICE_AFTER)
            col_dup,_        = ensure_column(ws, COL_DUP_JULY)
            col_manual,_     = ensure_column(ws, COL_MANUAL)
            col_approval,_   = ensure_column(ws, COL_APPROVAL)
            col_notes,_      = ensure_column(ws, COL_NOTES)
            col_total_pay,_  = ensure_column(ws, COL_TOTAL_PAY)
            col_diff,_       = ensure_column(ws, COL_DIFF_ROW)

            added_extra = []
            for name in EXTRA_COLUMNS:
                _, added = ensure_column(ws, name)
                if added:
                    added_extra.append(name)

            # locate existing "סה\"כ" column (do NOT create if missing)
            col_total = find_col(ws, COL_TOTAL)
            if col_total is None:
                st.warning(f'[{target_sheet}] לא נמצאה עמודה "סה\"כ". הזרקה ל"פער לפי שורה" תדלג (נדרש מקור חיסור).')

            # Letters
            L_src      = get_column_letter(col_purchase_src)
            L_rakhash  = get_column_letter(col_rakhash)
            L_makat    = get_column_letter(col_makat)
            L_clean    = get_column_letter(col_clean)
            L_order    = get_column_letter(col_order)
            L_qty      = get_column_letter(col_qty)
            L_qtychk   = get_column_letter(col_qty_check)
            L_price    = get_column_letter(col_price)
            L_manual   = get_column_letter(col_manual)
            L_approval = get_column_letter(col_approval)
            L_totalpay = get_column_letter(col_total_pay)
            L_diffcol  = get_column_letter(col_diff)
            L_totalcol = get_column_letter(col_total) if col_total else None

            max_row = ws.max_row
            cnt_rakhash = cnt_clean = cnt_order = cnt_qtychk = 0
            cnt_approval = cnt_totalpay = cnt_diff = 0

            # 4) Base formulas (full scan to create order key etc. like before)
            for r in range(2, max_row + 1):
                src_val = ws[f"{L_src}{r}"].value
                if src_val is not None and str(src_val).strip() != "":
                    ws[f"{L_rakhash}{r}"] = f"={L_src}{r}*1"
                    cnt_rakhash += 1

            for r in range(2, max_row + 1):
                v = ws[f"{L_makat}{r}"].value
                if v is not None and str(v).strip() != "":
                    ws[f"{L_clean}{r}"] = f"=LEFT({L_makat}{r},7)"
                    cnt_clean += 1

            for r in range(2, max_row + 1):
                v1 = ws[f"{L_rakhash}{r}"].value
                v2 = ws[f"{L_clean}{r}"].value
                if v1 is not None and v2 is not None and str(v1).strip() != "" and str(v2).strip() != "":
                    ws[f"{L_order}{r}"] = f"={L_rakhash}{r}&{L_clean}{r}"
                    cnt_order += 1

            # stop line determined by last non-empty in "הזמנות לבדיקה"
            end_row_orders = last_nonempty_row(ws, L_order, start_row=2)

            # 4.4 בדיקת כמות – VLOOKUP עם MATCH("סופק") על שורת הכותרות של הגיליון שהועתק
            for r in range(2, end_row_orders + 1):
                order_val = ws[f"{L_order}{r}"].value
                if order_val is not None and str(order_val).strip() != "":
                    # ✅ נרמול שמות הגיליונות למניעת בעיות של רווחים או תווים נסתרים
                    sheet_key = main_sheet_name.strip()
                    sheet_1_key = REQUIRED_MAIN_SHEET_1.strip()
                    sheet_2_key = REQUIRED_MAIN_SHEET_2.strip()

                    if sheet_key == sheet_1_key:  # "הובלה לבית לקוח"
                        # תנאי מיוחד: אם הכמות < 3 -> "תקין", אחרת "בדיקת כמות"
                        formula_qty = (
                            '=IFNA('
                            'IF(VLOOKUP({order},\'{internal}\'!$A:$O,'
                            'MATCH("סופק",\'{internal}\'!$A$1:$O$1,0),0)='
                            '\'{main}\'!{qty},'
                            'IF(\'{main}\'!{qty}<3,"תקין","בדיקת כמות"),'
                            '"נדרשת בדיקה"),'
                            '"נדרשת בדיקה")'
                        ).format(
                            order=f"{L_order}{r}",
                            internal=REQUIRED_INTERNAL_SHEET,
                            main=main_sheet_name,  # שומר את שם הגיליון האמיתי, גם אם כולל רווח
                            qty=f"{L_qty}{r}",
                        )

                    elif sheet_key == sheet_2_key:  # "הובלה לסוחר"
                        # תנאי פשוט: שוויון -> "תקין", אחרת "נדרשת בדיקה"
                        formula_qty = (
                            '=IFNA('
                            'IF(VLOOKUP({order},\'{internal}\'!$A:$O,'
                            'MATCH("סופק",\'{internal}\'!$A$1:$O$1,0),0)='
                            '{qty},'
                            '"תקין",'
                            '"נדרשת בדיקה"),'
                            '"נדרשת בדיקה")'
                        ).format(
                            order=f"{L_order}{r}",
                            internal=REQUIRED_INTERNAL_SHEET,
                            qty=f"{L_qty}{r}",
                        )

                    else:
                        # ברירת מחדל בטוחה (עם MATCH)
                        formula_qty = (
                            '=IFNA('
                            'IF(VLOOKUP({order},\'{internal}\'!$A:$O,'
                            'MATCH("סופק",\'{internal}\'!$A$1:$O$1,0),0)='
                            '{qty},'
                            '"תקין",'
                            '"נדרשת בדיקה"),'
                            '"נדרשת בדיקה")'
                        ).format(
                            order=f"{L_order}{r}",
                            internal=REQUIRED_INTERNAL_SHEET,
                            qty=f"{L_qty}{r}",
                        )

                    ws[f"{L_qtychk}{r}"] = formula_qty
                    cnt_qtychk += 1



            # 4.5 אישור סופי
            for r in range(2, end_row_orders + 1):
                formula_approval = (
                    '=IF(OR(AND({manual}{r}=0,{qtychk}{r}="תקין"),{manual}{r}="מאושר"),"מאושר","לא מאושר")'
                ).format(manual=L_manual, qtychk=L_qtychk, r=r)
                ws[f"{L_approval}{r}"] = formula_approval
                cnt_approval += 1

            # 4.6 סה"כ לתשלום
            for r in range(2, end_row_orders + 1):
                formula_total = (
                    '=IF({approval}{r}="מאושר",{price}{r}*ABS({qty}{r}),0)'
                ).format(approval=L_approval, price=L_price, qty=L_qty, r=r)
                ws[f"{L_totalpay}{r}"] = formula_total
                cnt_totalpay += 1

            # 4.7 פער לפי שורה = סה"כ - סה"כ לתשלום
            if L_totalcol:
                for r in range(2, end_row_orders + 1):
                    ws[f"{L_diffcol}{r}"] = f"={L_totalcol}{r}-{L_totalpay}{r}"
                    cnt_diff += 1

            added_msg = (" | נוספו עמודות: " + ", ".join(added_extra)) if added_extra else ""
            overall_msgs.append(
                f"✅ [{target_sheet}] הזרקות עד שורה {end_row_orders}. נוסחאות – רכש ({cnt_rakhash}), מק\"ט ללא פגומים ({cnt_clean}), הזמנות לבדיקה ({cnt_order}), בדיקת כמות/MATCH ({cnt_qtychk}), אישור סופי ({cnt_approval}), סה\"כ לתשלום ({cnt_totalpay}), פער לפי שורה ({cnt_diff}).{added_msg}"
            )

        # 5) Save and offer download
        out = BytesIO()
        wb.save(out)
        out.seek(0)

        for msg in overall_msgs:
            st.success(msg)

        st.download_button("📥 הורדת הדוח המעודכן",
            data=out,
            file_name="Activity_Unified_All_Sheets_With_ROWDIFF_and_MATCH.py.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error("שגיאה בעיבוד הקבצים.")
        st.exception(e)
else:
    st.info("נא להעלות את שני הקבצים (.xlsx).")




# ==== FOOTER (Single Instance) ====
st.markdown("""
<div class="footer">
  <div><b>נוצר ע״י:</b> אברהם מועלם, מנהל תפעול ומהנדס אוטומציות · <b>תאריך יצירה:</b> 24.10.2025</div>
  <div>© כל הזכויות שמורות 🧩</div>
  <div style="margin-top:4px;">טיפ: כדי למנוע שגיאות מבנה — העלו תמיד גרסה עדכנית של קובץ ‘הובלות אלכל כללי’.</div>
            
</div>
""", unsafe_allow_html=True)
# ==== END FOOTER ====
